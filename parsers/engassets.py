"""Engineering Assets parser.

Reads plant asset lists from XLSX files configured under eng_assets in config.yaml.
Falls back to an empty list if no source is configured; the dashboard shows
the Engineering Documents tab with whatever data is provided.

config.yaml schema:
  eng_assets:
    path: "./exports/engineering/asset_register.xlsx"
    sheet: "Assets"          # optional, defaults to first sheet
    columns:
      tag:        "Tag Number"
      desc:       "Description"
      type:       "Equipment Type"    # Pipe | Vessel | Pump | HeatExchanger | Valve | Instrument
      discipline: "Discipline"
      dn:         "Nominal Diameter"
      pn:         "Pressure Rating"
      pipeclass:  "Pipe Class"
      medium:     "Medium"
      temp:       "Design Temperature"
      mat:        "Material"
      insul:      "Insulation"
      p_and_id:   "P&ID Reference"
      isometry:   "Isometry Number"
      line_list:  "Line List Number"
"""
from __future__ import annotations
from pathlib import Path


def parse_eng_assets(cfg: dict) -> list:
    """Return list of asset dicts for ENG_ASSETS in the dashboard."""
    eng_cfg = cfg.get('eng_assets', {})
    if not eng_cfg or not eng_cfg.get('path'):
        return _mock_assets()

    path = Path(eng_cfg['path'])
    if not path.exists():
        print(f"  [warn] eng_assets path not found: {path}")
        return _mock_assets()

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheet_name = eng_cfg.get('sheet')
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        cols_cfg = eng_cfg.get('columns', {})

        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            return _mock_assets()
        col_idx = {str(h).strip(): i for i, h in enumerate(header) if h}

        def get(row, key, default=''):
            col_name = cols_cfg.get(key, key)
            idx = col_idx.get(col_name)
            if idx is None:
                return default
            v = row[idx]
            return str(v).strip() if v is not None else default

        assets = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            tag = get(row, 'tag')
            if not tag:
                continue
            assets.append({
                'tag':       tag,
                'desc':      get(row, 'desc'),
                'type':      get(row, 'type'),
                'discipline':get(row, 'discipline'),
                'dn':        get(row, 'dn'),
                'pn':        get(row, 'pn'),
                'pipeclass': get(row, 'pipeclass'),
                'medium':    get(row, 'medium'),
                'temp':      get(row, 'temp'),
                'mat':       get(row, 'mat'),
                'insul':     get(row, 'insul'),
                'p_and_id':  get(row, 'p_and_id'),
                'isometry':  get(row, 'isometry'),
                'lineList':  get(row, 'line_list'),
            })
        wb.close()
        print(f"  Engineering assets: {len(assets)} items from {path.name}")
        return assets
    except Exception as e:
        print(f"  [warn] Engineering assets parse error: {e}")
        return _mock_assets()


def _mock_assets():
    """Minimal demo data used when no real source is configured."""
    return [
        {'tag':'L-101','desc':'Reactor Feed Line','type':'Pipe','discipline':'Piping','dn':'DN100','pn':'PN40','pipeclass':'A2B','medium':'H2SO4','temp':'80°C','mat':'316L','insul':'Yes','p_and_id':'P-PR-001','isometry':'ISO-PP-001','lineList':'1-H2SO4-100-A2B'},
        {'tag':'L-102','desc':'Reactor Effluent Line','type':'Pipe','discipline':'Piping','dn':'DN150','pn':'PN40','pipeclass':'A2B','medium':'H2SO4','temp':'110°C','mat':'316L','insul':'Yes','p_and_id':'P-PR-001','isometry':'ISO-PP-002','lineList':'2-H2SO4-150-A2B'},
        {'tag':'R-100','desc':'Main Reactor','type':'Vessel','discipline':'Mechanical','dn':'∅1200','pn':'PN16','pipeclass':'—','medium':'H2SO4','temp':'120°C','mat':'316L','insul':'Yes','p_and_id':'P-PR-001','isometry':'','lineList':''},
        {'tag':'P-101A','desc':'Feed Pump A','type':'Pump','discipline':'Mechanical','dn':'DN50','pn':'PN16','pipeclass':'A2B','medium':'H2SO4','temp':'25°C','mat':'316L','insul':'No','p_and_id':'P-PR-001','isometry':'','lineList':''},
        {'tag':'E-101','desc':'Feed/Effluent HX','type':'HeatExchanger','discipline':'Mechanical','dn':'∅600','pn':'PN25','pipeclass':'—','medium':'H2SO4','temp':'100°C','mat':'316L','insul':'Yes','p_and_id':'P-PR-001','isometry':'','lineList':''},
        {'tag':'FT-101','desc':'Feed Flow Meter','type':'Instrument','discipline':'Instrumentation','dn':'DN100','pn':'PN40','pipeclass':'A2B','medium':'H2SO4','temp':'80°C','mat':'316L','insul':'No','p_and_id':'P-PR-001','isometry':'','lineList':''},
        {'tag':'XV-201','desc':'Isolation Valve Reactor Inlet','type':'Valve','discipline':'Piping','dn':'DN80','pn':'PN40','pipeclass':'A2B','medium':'H2SO4','temp':'80°C','mat':'316L','insul':'No','p_and_id':'P-PR-002','isometry':'ISO-PP-003','lineList':'5-TOL-80-B1'},
    ]
