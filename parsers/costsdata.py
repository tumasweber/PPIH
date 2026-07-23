"""Costs data parser.

Supports XLSX file exports from SAP, Primavera, or any ERP system.

config.yaml schema:
  costs:
    currency: "CHF"
    source: "xlsx"          # "xlsx" | "mock"

    actuals:
      path:  "./exports/costs/actual_costs.xlsx"
      sheet: "Actuals"
      columns:
        wbs:       "WBS Element"
        desc:      "Description"
        actual:    "Actual Costs"
        committed: "Committed"
        budget:    "Budget"

    scurve:
      path:  "./exports/costs/scurve.xlsx"
      sheet: "S-Curve"
      columns:
        date:     "Date"
        actual:   "Actual Cumulative"
        planned:  "Planned Cumulative"
        forecast: "Forecast"

Output shape (matches existing JS):
  {
    currency, budget_total, actual_total, forecast_total, committed,
    items: [{wbs, desc, budget, actual, committed}],
    by_discipline: [{disc, budget, actual}],
    scurve: {labels, budget, actual, forecast}
  }
"""
from __future__ import annotations
from pathlib import Path


def parse_costs(cfg: dict) -> dict | None:
    costs_cfg = cfg.get('costs', {})
    currency = costs_cfg.get('currency', 'CHF')
    if not costs_cfg or costs_cfg.get('source', 'mock') == 'mock':
        return _mock_costs(currency)

    items  = _parse_actuals(costs_cfg.get('actuals', {}))
    scurve = _parse_scurve(costs_cfg.get('scurve', {}))
    if not items:
        return _mock_costs(currency)

    return _build_output(currency, items, scurve)


def _parse_actuals(act_cfg: dict) -> list:
    if not act_cfg or not act_cfg.get('path'):
        return []
    path = Path(act_cfg['path'])
    if not path.exists():
        print(f"  [warn] Costs actuals not found: {path}")
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb[act_cfg['sheet']] if act_cfg.get('sheet') in wb.sheetnames else wb.active
        cc = act_cfg.get('columns', {})
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header: return []
        ci = {str(h).strip(): i for i, h in enumerate(header) if h}
        def g(row, k, d=0):
            idx = ci.get(cc.get(k, k))
            v = row[idx] if idx is not None and row[idx] is not None else d
            return float(v) if isinstance(v, (int, float)) else d
        def gs(row, k, d=''):
            idx = ci.get(cc.get(k, k))
            v = row[idx] if idx is not None else None
            return str(v).strip() if v is not None else d
        items = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            wbs = gs(row, 'wbs')
            if not wbs: continue
            items.append({'wbs': wbs, 'desc': gs(row,'desc'),
                          'budget': g(row,'budget'), 'actual': g(row,'actual'),
                          'committed': g(row,'committed')})
        wb.close()
        return items
    except Exception as e:
        print(f"  [warn] Costs actuals parse error: {e}")
        return []


def _parse_scurve(sc_cfg: dict) -> dict:
    empty = {'labels': [], 'budget': [], 'actual': [], 'forecast': []}
    if not sc_cfg or not sc_cfg.get('path'):
        return empty
    path = Path(sc_cfg['path'])
    if not path.exists():
        return empty
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb[sc_cfg['sheet']] if sc_cfg.get('sheet') in wb.sheetnames else wb.active
        cc = sc_cfg.get('columns', {})
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header: return empty
        ci = {str(h).strip(): i for i, h in enumerate(header) if h}
        def g(row, k):
            idx = ci.get(cc.get(k, k))
            v = row[idx] if idx is not None else None
            return float(v) if isinstance(v, (int, float)) else None
        labels, budget, actual, forecast = [], [], [], []
        for row in ws.iter_rows(min_row=2, values_only=True):
            idx = ci.get(cc.get('date', 'Date'))
            date = row[idx] if idx is not None else None
            if not date: continue
            labels.append(date.strftime('%Y-%m') if hasattr(date,'strftime') else str(date)[:7])
            budget.append(g(row, 'planned'))
            actual.append(g(row, 'actual'))
            forecast.append(g(row, 'forecast'))
        wb.close()
        return {'labels': labels, 'budget': budget, 'actual': actual, 'forecast': forecast}
    except Exception as e:
        print(f"  [warn] S-curve parse error: {e}")
        return empty


def _build_output(currency, items, scurve):
    total_b = sum(i['budget']    for i in items)
    total_a = sum(i['actual']    for i in items)
    total_c = sum(i['committed'] for i in items)
    total_f = total_a + total_c  # simple forecast
    # Group by first word of desc as discipline
    disc_map = {}
    for i in items:
        disc = i['wbs'].split('.')[0] if '.' in i['wbs'] else i['desc'].split()[0] if i['desc'] else i['wbs']
        disc_map.setdefault(disc, {'budget': 0, 'actual': 0})
        disc_map[disc]['budget'] += i['budget']
        disc_map[disc]['actual'] += i['actual']
    by_disc = [{'disc': k, 'budget': v['budget'], 'actual': v['actual']}
               for k, v in disc_map.items()]
    print(f"  Costs: {len(items)} WBS items, budget {currency} {total_b:,.0f}")
    return {'currency': currency, 'budget_total': total_b, 'actual_total': total_a,
            'forecast_total': total_f, 'committed': total_c,
            'items': items, 'by_discipline': by_disc, 'scurve': scurve}


def _mock_costs(currency='CHF'):
    items = [
        {'wbs':'1.1','desc':'Engineering',     'budget':480000,  'actual':312000,  'committed':48000},
        {'wbs':'1.2','desc':'Procurement',     'budget':1200000, 'actual':680000,  'committed':320000},
        {'wbs':'1.3','desc':'Civil/Structural','budget':380000,  'actual':95000,   'committed':180000},
        {'wbs':'1.4','desc':'Mechanical',      'budget':920000,  'actual':410000,  'committed':280000},
        {'wbs':'1.5','desc':'Piping',          'budget':640000,  'actual':285000,  'committed':210000},
        {'wbs':'1.6','desc':'Electrical/I&A',  'budget':520000,  'actual':168000,  'committed':95000},
        {'wbs':'1.7','desc':'Commissioning',   'budget':160000,  'actual':0,       'committed':0},
    ]
    total_b = sum(i['budget']    for i in items)
    total_a = sum(i['actual']    for i in items)
    total_c = sum(i['committed'] for i in items)
    by_disc = [{'disc':i['desc'],'budget':i['budget'],'actual':i['actual']} for i in items]
    scurve = {
        'labels':   ['2025-01','2025-02','2025-03','2025-04','2025-05','2025-06',
                     '2025-07','2025-08','2025-09','2025-10','2025-11','2025-12'],
        'budget':   [80000,160000,320000,520000,780000,1050000,
                     1380000,1700000,1950000,2200000,2380000,total_b],
        'actual':   [0,42000,128000,310000,580000,total_a,
                     None,None,None,None,None,None],
        'forecast': [None,None,None,None,None,total_a,
                     920000,1280000,1650000,1980000,2280000,int(total_b*1.05)],
    }
    return {'currency': currency, 'budget_total': total_b, 'actual_total': total_a,
            'forecast_total': int(total_b*1.05), 'committed': total_c,
            'items': items, 'by_discipline': by_disc, 'scurve': scurve}
