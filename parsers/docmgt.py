"""Document management parser (file import + SharePoint live query)."""
from __future__ import annotations
import os, re
try:
    import openpyxl
except ImportError:
    openpyxl = None

def apply_filters(issues, flt):
    def matches(issue):
        for field in ("status", "type", "priority"):
            allowed = flt.get(field, [])
            if allowed and issue.get(field) not in allowed:
                return False
        inc = flt.get("label_include", [])
        exc = flt.get("label_exclude", [])
        lbls = set(issue.get("labels", []))
        if inc and not lbls.intersection(inc):
            return False
        if exc and lbls.intersection(exc):
            return False
        return True
    return [i for i in issues if matches(i)]


# ─────────────────────────────────────────────────────────────────────────────
#  HTML TEMPLATE
# ─────────────────────────────────────────────────────────────────────────────


# ══════════════════════════════════════════════════════════════════
# Document Management Parser
# ══════════════════════════════════════════════════════════════════

def _doc_norm_status(raw, status_map):
    """Normalise a raw status string using the configured map."""
    if not raw:
        return "Missing"
    raw = str(raw).strip()
    return status_map.get(raw, raw)

def _doc_strip_ext(name):
    """Strip file extension and revision suffix for doc number matching.
    E.g. 'P-PR-001_Rev-B.pdf' → 'P-PR-001'"""
    import re
    name = str(name or '').strip()
    # Remove extension
    name = re.sub(r'\.[a-zA-Z]{2,5}$', '', name)
    # Remove common revision suffixes: _Rev-B, _RevB, _r02, -B, etc.
    name = re.sub(r'[_\-](?:Rev|REV|rev)?[_\-]?[A-Z0-9]{1,3}$', '', name)
    return name.strip()

def _doc_col(row, col_map, key, ws_headers):
    """Read a cell value by configured column name."""
    col_name = col_map.get(key, '')
    if not col_name or col_name not in ws_headers:
        return ''
    idx = ws_headers[col_name]
    val = row[idx]
    return str(val).strip() if val is not None else ''

def _fetch_sharepoint_live(sp_cfg: dict):
    """
    Fetch document list from SharePoint via Microsoft Graph API.
    Returns (documents: list[dict], diag: dict).
    diag contains step-by-step connection diagnostics for the debug panel.
    Modes: 'folder' (scan library/subfolder) | 'file' (single xlsx/csv)
    """
    import re as _re

    diag = {
        'query_mode':  sp_cfg.get('query_mode', 'folder'),
        'site_url':    sp_cfg.get('site_url', ''),
        'library':     sp_cfg.get('library', ''),
        'steps':       [],   # list of {step, ok, detail}
        'error':       None,
    }

    def step(name, ok, detail=''):
        icon = '✓' if ok else '✗'
        print(f'  [DOC/SP] {icon} {name}: {detail}' if detail else f'  [DOC/SP] {icon} {name}')
        diag['steps'].append({'step': name, 'ok': ok, 'detail': detail})
        if not ok:
            diag['error'] = f'{name}: {detail}' if detail else name

    query_mode    = sp_cfg.get('query_mode', 'folder')
    site_url      = sp_cfg.get('site_url', '').rstrip('/')
    tenant_id     = sp_cfg.get('tenant_id', '')
    client_id     = sp_cfg.get('client_id', '')
    client_secret = sp_cfg.get('client_secret', '')

    # ── Credential check ─────────────────────────────────────────────
    missing = [k for k,v in [('site_url',site_url),('tenant_id',tenant_id),
                               ('client_id',client_id),('client_secret',client_secret)] if not v]
    if missing:
        step('Credentials', False, f'missing: {", ".join(missing)}')
        return [], diag
    step('Credentials', True, f'site={site_url}')

    try:
        import requests
    except ImportError:
        step('Import requests', False, 'pip install requests')
        return [], diag
    step('Import requests', True)

    # ── 1. Get access token ──────────────────────────────────────────
    try:
        token_resp = requests.post(
            f'https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token',
            data={'grant_type':'client_credentials','client_id':client_id,
                  'client_secret':client_secret,
                  'scope':'https://graph.microsoft.com/.default'},
            timeout=30)
        if token_resp.status_code != 200:
            err = token_resp.json().get('error_description', token_resp.text[:120])
            step('OAuth2 token', False, f'HTTP {token_resp.status_code} — {err}')
            return [], diag
        token = token_resp.json().get('access_token', '')
        step('OAuth2 token', True, f'tenant {tenant_id[:8]}…')
    except Exception as e:
        step('OAuth2 token', False, str(e)[:120])
        return [], diag

    hdrs = {'Authorization': f'Bearer {token}', 'Accept': 'application/json'}

    # ── 2. Resolve site ID ───────────────────────────────────────────
    try:
        from urllib.parse import urlparse
        p = urlparse(site_url)
        site_api  = f'https://graph.microsoft.com/v1.0/sites/{p.netloc}:/{p.path.lstrip("/")}'
        site_resp = requests.get(site_api, headers=hdrs, timeout=30)
        if site_resp.status_code != 200:
            step('Resolve site', False, f'HTTP {site_resp.status_code} — {site_resp.text[:100]}')
            return [], diag
        site_id = site_resp.json().get('id', '')
        site_name = site_resp.json().get('displayName', site_url)
        step('Resolve site', True, f'"{site_name}" → {site_id[:20]}…')
    except Exception as e:
        step('Resolve site', False, str(e)[:120])
        return [], diag

    # ── 3. Resolve drive (library) ───────────────────────────────────
    try:
        library   = sp_cfg.get('library', '')
        drives_r  = requests.get(
            f'https://graph.microsoft.com/v1.0/sites/{site_id}/drives',
            headers=hdrs, timeout=30)
        drives    = drives_r.json().get('value', [])
        drive_names = [d.get('name','') for d in drives]
        drive = next((d for d in drives
                      if not library or d.get('name','').lower() == library.lower()),
                     drives[0] if drives else None)
        if not drive:
            step('Resolve library', False,
                 f'"{library}" not found. Available: {", ".join(drive_names) or "(none)"}')
            return [], diag
        drive_id = drive['id']
        step('Resolve library', True,
             f'"{drive.get("name")}" ({len(drives)} libraries available: {", ".join(drive_names)})')
    except Exception as e:
        step('Resolve library', False, str(e)[:120])
        return [], diag

    # ════════════════════════════════════════════════════════════════
    # MODE A — Single file
    # ════════════════════════════════════════════════════════════════
    if query_mode == 'file':
        file_path = sp_cfg.get('file_path', '').strip('/')
        if not file_path:
            step('Single file', False, 'file_path not configured')
            return [], diag

        try:
            content_url = (f'https://graph.microsoft.com/v1.0'
                           f'/drives/{drive_id}/root:/{file_path}:/content')
            resp = requests.get(content_url, headers=hdrs, timeout=60)
            if resp.status_code != 200:
                step('Download file', False,
                     f'HTTP {resp.status_code} — {file_path} '
                     f'({resp.text[:80]})')
                return [], diag
            step('Download file', True, f'{file_path} ({len(resp.content)//1024} KB)')
        except Exception as e:
            step('Download file', False, str(e)[:120])
            return [], diag

        try:
            import openpyxl, io as _io
            wb = openpyxl.load_workbook(_io.BytesIO(resp.content),
                                        read_only=True, data_only=True)
            sheet_name = sp_cfg.get('sheet', '')
            ws = (wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames
                  else wb.active)
            rows = list(ws.rows)
            if not rows:
                step('Parse file', False, 'Sheet is empty')
                return [], diag

            headers = {str(c.value or '').strip(): i for i, c in enumerate(rows[0])}
            doc_col    = sp_cfg.get('doc_number_column', '')
            title_col  = sp_cfg.get('title_column', 'Title')
            status_col = sp_cfg.get('status_column', 'Status')
            rev_col    = sp_cfg.get('revision_column', 'Revision')
            date_col   = sp_cfg.get('date_column', 'Date')

            if doc_col not in headers:
                step('Parse file', False,
                     f'Doc number column "{doc_col}" not found. '
                     f'Available: {", ".join(list(headers)[:8])}')
                return [], diag

            def _gv(row, col):
                i = headers.get(col, -1)
                return str(row[i].value or '').strip() if 0 <= i < len(row) else ''

            smap      = sp_cfg.get('status_map', {})
            documents = []
            for row in rows[1:]:
                dn = _gv(row, doc_col)
                if not dn:
                    continue
                raw_st = _gv(row, status_col)
                documents.append({
                    'doc_number':  dn,
                    'title':       _gv(row, title_col),
                    'status':      smap.get(raw_st, raw_st),
                    'revision':    _gv(row, rev_col),
                    'date':        _gv(row, date_col),
                    'discipline':  _gv(row, sp_cfg.get('discipline_column', 'Discipline')),
                    'modified_by': '',
                    'path':        f'{site_url}/{file_path}',
                })
            step('Parse file', True,
                 f'{len(documents)} docs from sheet "{ws.title}" '
                 f'({len(rows)-1} data rows, cols: {", ".join(list(headers)[:6])}…)')
            return documents, diag
        except Exception as e:
            step('Parse file', False, str(e)[:120])
            return [], diag

    # ════════════════════════════════════════════════════════════════
    # MODE B — Folder scan
    # ════════════════════════════════════════════════════════════════
    folder_path  = sp_cfg.get('folder_path', '')
    file_pattern = sp_cfg.get('file_pattern', r'\.(pdf|docx|xlsx)$')
    doc_num_pat  = sp_cfg.get('doc_number_pattern', '')
    status_col   = sp_cfg.get('status_column', 'ApprovalStatus')

    try:
        if folder_path:
            items_url = (f'https://graph.microsoft.com/v1.0'
                         f'/drives/{drive_id}/root:/{folder_path}:/children')
        else:
            items_url = (f'https://graph.microsoft.com/v1.0'
                         f'/drives/{drive_id}/root/children')
        # Test the URL first
        test_r = requests.get(items_url, headers=hdrs, timeout=30)
        if test_r.status_code != 200:
            step('List folder', False,
                 f'HTTP {test_r.status_code} — '
                 f'{"folder not found" if test_r.status_code == 404 else test_r.text[:80]}')
            return [], diag
        n_items = len(test_r.json().get('value', []))
        step('List folder', True,
             f'{folder_path or "(root)"} → {n_items} items (first page)')
    except Exception as e:
        step('List folder', False, str(e)[:120])
        return [], diag

    file_pat = _re.compile(file_pattern, _re.IGNORECASE) if file_pattern else None
    doc_pat  = _re.compile(doc_num_pat) if doc_num_pat else None
    smap     = sp_cfg.get('status_map', {})

    documents, next_url, n_pages = [], items_url, 0
    try:
        while next_url:
            resp  = requests.get(next_url, headers=hdrs, timeout=30).json()
            items = resp.get('value', [])
            n_pages += 1
            for item in items:
                name = item.get('name', '')
                if item.get('folder') or (file_pat and not file_pat.search(name)):
                    continue
                doc_number = ''
                if doc_pat:
                    m = doc_pat.search(name)
                    if m and m.groups():
                        doc_number = m.group(1)
                doc_number = doc_number or name
                fields     = item.get('listItem', {}).get('fields', {})
                raw_st     = str(fields.get(status_col, '') or '')
                documents.append({
                    'doc_number':  doc_number,
                    'title':       name,
                    'status':      smap.get(raw_st, raw_st),
                    'revision':    '',
                    'date':        item.get('lastModifiedDateTime', '')[:10],
                    'modified_by': (item.get('lastModifiedBy', {})
                                    .get('user', {}).get('displayName') or ''),
                    'path':        item.get('webUrl', ''),
                })
            next_url = resp.get('@odata.nextLink')
        step('Scan files', True,
             f'{len(documents)} matching files across {n_pages} page(s)'
             + (f' (pattern: {file_pattern})' if file_pattern else ''))
    except Exception as e:
        step('Scan files', False, str(e)[:120])
        return documents, diag

    return documents, diag


def parse_doc_management(cfg, config_dir=None):
    """
    Parse document management sources and cross-reference against master register.
    Returns a dict ready for JSON embedding.
    """
    dm_cfg = cfg.get('doc_management', {})
    if not dm_cfg.get('enabled', False):
        return None

    try:
        import openpyxl
    except ImportError:
        print("[warn] openpyxl not installed — skipping document management")
        return None

    def read_xlsx(path, sheet_name, col_map, status_map=None, strip_ext=False):
        """Read an xlsx sheet and return list of dicts using col_map."""
        if not path or not os.path.isfile(path):
            print(f"  [doc] file not found: {path}")
            return []
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            rows = list(ws.rows)
            if not rows:
                return []
            # Build header index
            headers = {str(cell.value or '').strip(): i for i, cell in enumerate(rows[0])}
            results = []
            for row in rows[1:]:
                vals = [cell.value for cell in row]
                def get(key):
                    cname = col_map.get(key, '')
                    if not cname or cname not in headers:
                        return ''
                    v = vals[headers[cname]] if headers[cname] < len(vals) else None
                    return str(v).strip() if v is not None else ''
                doc_num = get('doc_number')
                if not doc_num:
                    continue
                if strip_ext:
                    doc_num = _doc_strip_ext(doc_num)
                raw_status = get('status')
                status = _doc_norm_status(raw_status, status_map or {})
                results.append({
                    'doc_number':  doc_num,
                    'title':       get('title'),
                    'discipline':  get('discipline'),
                    'status':      status,
                    'revision':    get('revision'),
                    'date':        get('date'),
                    'modified_by': get('modified_by'),
                    'model_file':  get('model_file') if 'model_file' in col_map else '',
                })
            wb.close()
            return results
        except Exception as e:
            print(f"  [doc] error reading {path}: {e}")
            return []

    # ── Read master register ──────────────────────────────────────
    mr_cfg  = dm_cfg.get('master_register', {})
    mr_path = mr_cfg.get('path', '')
    if mr_path and not os.path.isabs(mr_path):
        base = config_dir or os.path.dirname(os.path.abspath(__file__))
        mr_path = os.path.join(base, mr_path)

    master_docs = {}
    if mr_path and os.path.isfile(mr_path):
        try:
            import openpyxl
            wb  = openpyxl.load_workbook(mr_path, read_only=True, data_only=True)
            ws  = wb[mr_cfg.get('sheet', '')] if mr_cfg.get('sheet','') in wb.sheetnames else wb.active
            rows = list(ws.rows)
            mc  = mr_cfg.get('columns', {})
            hdrs = {str(c.value or '').strip(): i for i, c in enumerate(rows[0])}
            def mget(row_vals, key):
                cname = mc.get(key, '')
                if not cname or cname not in hdrs: return ''
                v = row_vals[hdrs[cname]] if hdrs[cname] < len(row_vals) else None
                return str(v).strip() if v is not None else ''
            for row in rows[1:]:
                vals = [c.value for c in row]
                dn = mget(vals, 'doc_number')
                if not dn: continue
                master_docs[dn] = {
                    'doc_number':      dn,
                    'title':           mget(vals, 'title'),
                    'discipline':      mget(vals, 'discipline'),
                    'asset_type':      mget(vals, 'asset_type'),
                    'required_status': mget(vals, 'required_status'),
                    'required_date':   mget(vals, 'required_date'),
                }
            wb.close()
            print(f"  [doc] master register: {len(master_docs)} documents")
        except Exception as e:
            print(f"  [doc] error reading master register: {e}")
    else:
        print(f"  [doc] master register not found: {mr_path}")

    # ── Read each source ──────────────────────────────────────────
    sources_cfg = dm_cfg.get('sources', {})
    _sp_diags = {}  # SharePoint connection diagnostics per source key
    source_results = {}  # source_key → list of doc dicts

    for src_key, src_cfg in sources_cfg.items():
        if not src_cfg.get('enabled', True):
            continue
        src_type = src_cfg.get('source_type', 'file')
        src_path = src_cfg.get('path', '')
        if src_path and not os.path.isabs(src_path):
            base = config_dir or os.path.dirname(os.path.abspath(__file__))
            src_path = os.path.join(base, src_path)
        label    = src_cfg.get('label', src_key)
        cols     = src_cfg.get('columns', {})
        smap     = src_cfg.get('status_map', {})
        sheet    = src_cfg.get('sheet', '')
        strip    = src_key == 'sharepoint'

        if src_type == 'sharepoint_live':
            sp_cfg        = src_cfg.get('sharepoint', {})
            raw_docs, _sp_diag = _fetch_sharepoint_live(sp_cfg)
            for d in raw_docs:
                if smap and d['status'] in smap:
                    d['status'] = smap[d['status']]
            docs = raw_docs
            # Store diagnostics for the debug panel
            _sp_diags[src_key] = _sp_diag
        else:
            docs = read_xlsx(src_path, sheet, cols, smap, strip_ext=strip)

        source_results[src_key] = {'label': label, 'docs': docs}
        print(f"  [doc] {label}: {len(docs)} documents")

    # ── Cross-reference: build unified document list ──────────────
    # Start from master register, enrich with source data
    status_priority = {'Approved': 4, 'Signed': 3, 'In Review': 2, 'Missing': 1, 'Rejected': 0}

    unified = []
    for dn, master in master_docs.items():
        doc_entry = dict(master)
        doc_entry['sources'] = {}

        # Find this doc in each source
        for src_key, src_data in source_results.items():
            match = next((d for d in src_data['docs'] if d['doc_number'] == dn), None)
            if match:
                doc_entry['sources'][src_key] = {
                    'status':      match['status'],
                    'revision':    match['revision'],
                    'date':        match['date'],
                    'modified_by': match['modified_by'],
                    'title':       match['title'],
                }

        # Derive overall status: worst status across all sources,
        # or "Missing" if not found in any source
        found_statuses = [s['status'] for s in doc_entry['sources'].values()]
        if not found_statuses:
            doc_entry['overall_status'] = 'Missing'
        else:
            # Use the lowest-priority status (e.g. if one source says In Review, use that)
            doc_entry['overall_status'] = min(
                found_statuses,
                key=lambda s: status_priority.get(s, 0)
            )

        # Check if overall status meets required
        req = master.get('required_status', '')
        actual_p = status_priority.get(doc_entry['overall_status'], 0)
        req_p    = status_priority.get(req, 0)
        doc_entry['meets_requirement'] = actual_p >= req_p

        unified.append(doc_entry)

    # Also add docs found in sources but NOT in master register
    all_master_nums = set(master_docs.keys())
    for src_key, src_data in source_results.items():
        for d in src_data['docs']:
            if d['doc_number'] not in all_master_nums:
                unified.append({
                    'doc_number':      d['doc_number'],
                    'title':           d.get('title', ''),
                    'discipline':      d.get('discipline', ''),
                    'asset_type':      d.get('asset_type', ''),
                    'required_status': '',
                    'required_date':   '',
                    'overall_status':  d.get('status', ''),
                    'meets_requirement': None,  # no requirement defined
                    'sources': {src_key: {
                        'status': d.get('status', ''), 'revision': d.get('revision', ''),
                        'date': d.get('date', ''), 'modified_by': d.get('modified_by', ''),
                        'title': d.get('title', ''),
                    }},
                })
                all_master_nums.add(d['doc_number'])

    # ── Summary stats ─────────────────────────────────────────────
    status_counts = {}
    for d in unified:
        s = d['overall_status']
        status_counts[s] = status_counts.get(s, 0) + 1

    asset_types = dm_cfg.get('asset_types', [])
    status_colors = dm_cfg.get('status_colors', {})

    print(f"  [doc] unified: {len(unified)} documents, status: {status_counts}")

    # Per-source debug metadata for the debug panel
    src_debug = {}
    for key, data in source_results.items():
        src_cfg = (dm_cfg.get('sources') or {}).get(key, {})
        src_debug[key] = {
            'label':       data.get('label', key),
            'doc_count':   len(data.get('docs', [])),
            'source_type': src_cfg.get('source_type', 'file'),
            'path':        src_cfg.get('path', '') if src_cfg.get('source_type') != 'sharepoint_live' else '',
            'sp_diag':     _sp_diags.get(key),   # None for non-live sources
        }

    return {
        'enabled':       True,
        'documents':     unified,
        'master_count':  len(master_docs),
        'status_counts': status_counts,
        'status_colors': status_colors,
        'asset_types':   asset_types,
        'sources':       src_debug,
        # Flat label map used by JS pill builder: {key: "Display Label"}
        'source_labels': {k: v['label'] for k, v in src_debug.items()},
    }

