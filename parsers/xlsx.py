"""Revizto Excel export (.xlsx) parser."""
from __future__ import annotations
import os, re, zipfile
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from .utils import _sanitise, _extract_discipline, _parse_date


# ─────────────────────────────────────────────────────────────────────────────
#  XLSX PARSER  (Revizto Excel export)
# ─────────────────────────────────────────────────────────────────────────────

def _sanitise(s):
    """
    Remove characters that would break JSON inside a <script> block:
    null bytes, lone surrogates, and other control characters except
    tab/newline/CR which json.dumps handles safely.
    Also strips the vertical tab and form-feed control chars common
    in copy-pasted text from Windows apps.
    """
    if not isinstance(s, str):
        return s
    # Remove null bytes and non-printable control chars (keep \t \n \r)
    return "".join(c for c in s if c == "\t" or c == "\n" or c == "\r"
                   or (ord(c) >= 32 and ord(c) != 127))


def _xlsx_val(row, idx, default=""):
    """Safely get cell value from a row tuple by 0-based index."""
    try:
        v = row[idx]
        if v is None:
            return default
        if hasattr(v, 'strftime'):          # datetime
            return v.strftime("%Y-%m-%dT%H:%M:%SZ")
        return _sanitise(str(v).strip())
    except IndexError:
        return default


def _xlsx_coords(raw):
    """
    Parse a Revizto coordinate string like '(40.45; -128.98; 19.29)'
    into a dict {'x': 40.45, 'y': -128.98, 'z': 19.29} or None.
    """
    if not raw:
        return None
    import re
    nums = re.findall(r'-?\d+\.?\d*', raw)
    if len(nums) >= 3:
        try:
            return {'x': float(nums[0]), 'y': float(nums[1]), 'z': float(nums[2])}
        except ValueError:
            pass
    return None


def col_letter_of_header(sheet_bytes, header_name):
    """
    Given raw sheet XML bytes, find the column letter of the cell in row 1
    whose shared-string value matches header_name.
    Returns e.g. 'C' or None.
    """
    try:
        import re as _re
        # Find cells in row 1 — look for the column with the Snapshot header
        # We parse the XML directly to avoid needing openpyxl here
        root = ET.fromstring(sheet_bytes)
        ns   = {'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        row1 = None
        for row in root.findall('.//x:row', ns):
            if row.get('r') == '1':
                row1 = row
                break
        if row1 is None:
            return None
        for cell in row1.findall('x:c', ns):
            ref = cell.get('r', '')
            v   = cell.find('x:v', ns)
            # The cell type 's' means shared string — we just match by position
            # Simpler: look for which column has the header by checking the ref letter
            # We'll match after building the col map in openpyxl — just return None here
            # and fall back to scanning all hyperlink columns
        return None  # let caller scan all hyperlinks
    except Exception:
        return None


def parse_xlsx_file(path, spatial_priority=None):
    """
    Parse a Revizto Excel export (.xlsx) and return a list of issue dicts
    using the same schema as parse_bcf_file so both can be merged.

    Revizto exports two sheets:
    - 'Raw Data'  — one row per issue, all fields (preferred)
    - 'Human-readable' — multi-row per issue (comments on extra rows)

    The 'Snapshot' cell (column C in Raw Data, column B in Human-readable)
    contains an external hyperlink to a Revizto S3 CDN URL for the issue
    snapshot image. This URL is extracted from the sheet's .rels file and
    stored as snapshot_url on each issue so the browser can load it directly
    when viewing the dashboard while authenticated with Revizto.
    """
    try:
        import openpyxl
    except ImportError:
        print("[error] openpyxl not installed. Run: pip install openpyxl")
        return []

    issues = []
    source  = os.path.basename(path)
    project = os.path.splitext(source)[0]

    # ── Extract snapshot URLs from the xlsx ZIP rels before openpyxl ────────
    # The Snapshot cell hyperlinks live in xl/worksheets/_rels/sheet*.xml.rels
    # Revizto puts the snapshot URL in column C (Raw Data) as an external hyperlink.
    # We build: row_number → URL, then attach the URL to each issue by row order.
    snapshot_url_by_row = {}   # data row index (1-based from header) → URL
    try:
        with zipfile.ZipFile(path, 'r') as z:
            wb_xml     = ET.fromstring(z.read('xl/workbook.xml'))
            wb_rels    = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
            wb_ns      = {'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            rels_ns    = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
            r_ns_full  = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

            # Find the file path for the "Raw Data" sheet
            rid_to_target = {
                r.get('Id'): r.get('Target','')
                for r in wb_rels.findall('r:Relationship', rels_ns)
                if 'worksheet' in r.get('Type','')
            }
            raw_sheet_file = None
            for sheet in wb_xml.findall('.//x:sheet', wb_ns):
                if sheet.get('name','') == 'Raw Data':
                    rid = sheet.get(f'{{{r_ns_full}}}id','')
                    if rid in rid_to_target:
                        raw_sheet_file = rid_to_target[rid]   # e.g. "worksheets/sheet2.xml"
                    break

            if raw_sheet_file:
                sheet_fname = raw_sheet_file.split('/')[-1]           # "sheet2.xml"
                rels_path   = f'xl/worksheets/_rels/{sheet_fname}.rels'
                sheet_path  = f'xl/{raw_sheet_file}'

                if rels_path in z.namelist() and sheet_path in z.namelist():
                    # Build rId → URL
                    rid_to_url = {
                        r.get('Id'): r.get('Target','')
                        for r in ET.fromstring(z.read(rels_path))
                                    .findall('r:Relationship', rels_ns)
                        if r.get('Target','').startswith('http')
                    }

                    # Parse hyperlinks in the sheet — map row number → URL
                    sheet_root = ET.fromstring(z.read(sheet_path))
                    sheet_ns2  = {'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                    import re as _re
                    for hl in sheet_root.findall('.//x:hyperlink', sheet_ns2):
                        ref = hl.get('ref','')          # e.g. "C2"
                        rid = hl.get(f'{{{r_ns_full}}}id','')
                        url = rid_to_url.get(rid,'')
                        if not url:
                            url = hl.get('location','') or hl.get('Target','')
                        if url and ref:
                            m = _re.match(r'[A-Z]+(\d+)$', ref)
                            if m:
                                row_num = int(m.group(1))
                                # Only store the first URL per row (the snapshot)
                                if row_num not in snapshot_url_by_row:
                                    snapshot_url_by_row[row_num] = url

    except Exception as e:
        print(f"  [warn] Could not extract snapshot URLs from {path}: {e}")

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"[error] Cannot open {path}: {e}")
        return []

    # Prefer 'Raw Data' sheet; fall back to first sheet
    ws = wb['Raw Data'] if 'Raw Data' in wb.sheetnames else wb.active

    # Build header → column-index map from row 1
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        print(f"  [warn] {path}: empty sheet")
        return []

    col = {str(h).strip(): i for i, h in enumerate(header_row) if h is not None}

    def _get(row, name, default=""):
        idx = col.get(name)
        if idx is None:
            return default
        return _xlsx_val(row, idx, default)

    # Parse comments from 'Human-readable' sheet (multi-row per issue)
    comments_by_guid = {}
    if 'Human-readable' in wb.sheetnames:
        hr = wb['Human-readable']
        hr_header = next(hr.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if hr_header:
            hc = {str(h).strip(): i for i, h in enumerate(hr_header) if h is not None}
            current_guid = None
            for row in hr.iter_rows(min_row=2, values_only=True):
                guid_val = _xlsx_val(row, hc.get('GUID', -1))
                if guid_val:
                    current_guid = guid_val
                if current_guid is None:
                    continue
                text   = _xlsx_val(row, hc.get('Comment', -1))
                author = _xlsx_val(row, hc.get('Comment Reporter', -1))
                date   = _xlsx_val(row, hc.get('Comment Date', -1))
                if text and text != 'Original Markup':
                    comments_by_guid.setdefault(current_guid, []).append({
                        'text':   text,
                        'author': author,
                        'date':   _parse_date(date) if date else '',
                    })

    now = datetime.now(timezone.utc)

    for data_row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            guid    = _get(row, 'GUID')
            if not guid:
                continue

            # Snapshot URL from rels (keyed by xlsx row number, 1-based)
            snapshot_url = snapshot_url_by_row.get(data_row_idx, "")

            title    = _get(row, 'Title') or _get(row, 'Stamp Title')
            status   = _get(row, 'Status')
            priority = _get(row, 'Priority')
            # Type: prefer workflow Type field if it's meaningful (not a stamp code/title),
            # otherwise use Stamp Abbreviation (e.g. "A040") which is always defined
            raw_type     = _get(row, 'Type')
            stamp_abbr   = _get(row, 'Stamp Abbreviation')
            stamp_cat    = _get(row, 'Stamp Category')
            # If Type column contains the stamp title (same as title) or is empty, use stamp_abbr
            if not raw_type or raw_type == title or raw_type == _get(row, 'Stamp Title'):
                ttype = stamp_abbr or stamp_cat or "Unknown"
            else:
                ttype = raw_type
            assigned_raw = _get(row, 'Assignee') or _get(row, 'Assigned To') or ''
            assignees_list_x = [a.strip() for a in assigned_raw.split(',') if a.strip()]
            assigned = ', '.join(assignees_list_x) if assignees_list_x else ''
            assigned = assigned  # keep for compat
            reporter = _get(row, 'Reporter')
            created  = _parse_date(_get(row, 'Created'))
            due      = _parse_date(_get(row, 'Deadline'))
            modified = _parse_date(_get(row, 'Last Updated'))

            # Spatial fields — the core reason for XLSX import
            level         = _get(row, 'Level')
            grid_location = _get(row, 'Grid Location')
            room          = _get(row, 'Room')
            space         = _get(row, 'Space')
            area          = _get(row, 'Area')
            zone          = _get(row, 'Zone')
            discipline_raw= _get(row, 'Discipline')   # Revizto native discipline
            coords_m      = _get(row, 'Coordinates (m)')
            coords_parsed = _xlsx_coords(coords_m)

            # Best available spatial label for grouping
            _sfp = spatial_priority or ["room","space","zone","area","level","grid_location"]
            _spatial_candidates = {
                "room": room, "space": space, "zone": zone,
                "area": area, "level": level, "grid_location": grid_location
            }
            spatial = next((v for f in _sfp
                            if (v := _spatial_candidates.get(f,""))), "") or ""

            # Discipline resolution priority:
            #   1. Stamp prefix from title (e.g. "H" from "H040_Duct clash")
            #      → matches discipline_map in config.yaml → resolved in JS
            #   2. Revizto native discipline_raw (full name, e.g. "HVAC")
            #      → stored as-is; JS disciplineName() will pass it through
            #      if it doesn't match a single-letter prefix
            #   3. Empty string (no data) — aggregation will skip it
            disc_prefix = _extract_discipline(title)
            discipline  = disc_prefix if disc_prefix else (discipline_raw.strip() if discipline_raw else "")

            # Tags → labels list
            tags_raw = _get(row, 'Tags')
            labels   = [t.strip() for t in tags_raw.split(',') if t.strip()] if tags_raw else []

            # Age and overdue
            age_days = ""
            overdue  = False
            if created:
                try:
                    c_dt = datetime.strptime(created, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)
                    age_days = (now - c_dt).days
                except Exception:
                    pass
            if due and status not in ("Closed", "Resolved", "Solved"):
                try:
                    d_dt = datetime.strptime(due, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)
                    overdue = now > d_dt
                except Exception:
                    pass

            comments_data = comments_by_guid.get(guid, [])

            issues.append({
                "guid":          guid,
                "title":         title or "(no title)",
                "status":        status or "Unknown",
                "type":          ttype  or "Unknown",
                "priority":      priority or "",
                "assigned":      assigned or "Unassigned",
                "description":   "",          # not in XLSX export
                "labels":        labels,
                "discipline":    discipline,                # prefix letter or raw name for JS map
                "discipline_raw": discipline_raw,              # Revizto full name
                # ── spatial fields (XLSX only) ──
                "level":         level,
                "grid_location": grid_location,
                "room":          room,
                "space":         space,
                "area":          area,
                "zone":          zone,
                "spatial":       spatial,      # best available label
                "coords_m":      coords_m,
                "coords":        coords_parsed,
                # ── standard fields ──
                "created":       created,
                "modified":      modified,
                "due":           due,
                "author":        reporter,
                "stage":         "",
                "age_days":      age_days,
                "overdue":       overdue,
                "comment_count": len(comments_data),
                "comments":      comments_data,
                "thumbnail":     "",           # XLSX has no embedded images
                "has_snapshot":  False,
                "snapshot_url":  snapshot_url, # Revizto S3 URL — loads in browser if authenticated
                "source_file":   source,
                        "project":       project,
                "bcf_version":   "xlsx",
            })
        except Exception as e:
            print(f"  [warn] Skipping row in {path}: {e}")

    wb.close()
    print(f"    -> {len(issues)} issues (XLSX, spatial data available)")
    return issues

