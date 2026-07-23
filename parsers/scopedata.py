"""Scope data parser.

Reads the project scope register from an XLSX file.
One row = one document type required for a system/discipline combination.

Output shape:
  {
    systems:  [{code, name, discipline, company, area}],
    doc_types: ['P&ID','Equipment List','Isometric',...],  # ordered unique list
    entries:  [{system_code, doc_type, discipline, company,
                required, doc_number, status_override}]
  }

config.yaml schema:
  scope:
    path:  "./exports/scope/scope_register.xlsx"
    sheet: "Scope"
    columns:
      system_code: "System Code"      # e.g. CW-001
      system_name: "System Name"      # e.g. Cooling Water
      discipline:  "Discipline"
      company:     "Company"
      area:        "Area"             # optional
      doc_type:    "Document Type"    # e.g. P&ID, Equipment List
      required:    "Required"         # YES/NO or 1/0 or X
      doc_number:  "Document Number"  # optional explicit link to DOC_DATA
"""
from __future__ import annotations
from pathlib import Path


def parse_scope(cfg: dict) -> dict:
    scope_cfg = cfg.get('scope', {})
    if not scope_cfg or not scope_cfg.get('path'):
        return _mock_scope()

    path = Path(scope_cfg['path'])
    if not path.exists():
        print(f"  [warn] Scope register not found: {path}")
        return _mock_scope()

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = (wb[scope_cfg['sheet']] if scope_cfg.get('sheet') in wb.sheetnames
              else wb.active)
        cc = scope_cfg.get('columns', {})

        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            return _mock_scope()
        ci = {str(h).strip(): i for i, h in enumerate(header) if h}

        def gs(row, key, default=''):
            col = cc.get(key, key)
            idx = ci.get(col)
            v = row[idx] if idx is not None and row[idx] is not None else default
            return str(v).strip() if v is not None else default

        def is_required(val):
            v = str(val or '').strip().upper()
            return v in ('YES', 'Y', 'X', '1', 'TRUE', 'OUI', 'JA')

        systems_seen = {}   # code → {name, discipline, company, area}
        entries = []
        doc_types_ordered = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            code = gs(row, 'system_code')
            if not code:
                continue
            name       = gs(row, 'system_name')
            discipline = gs(row, 'discipline')
            company    = gs(row, 'company')
            area       = gs(row, 'area')
            doc_type   = gs(row, 'doc_type')
            required   = is_required(gs(row, 'required', 'YES'))
            doc_number = gs(row, 'doc_number')

            if code not in systems_seen:
                systems_seen[code] = {'code': code, 'name': name,
                                      'discipline': discipline,
                                      'company': company, 'area': area}
            if doc_type and doc_type not in doc_types_ordered:
                doc_types_ordered.append(doc_type)

            entries.append({
                'system_code': code,
                'doc_type':    doc_type,
                'discipline':  discipline,
                'company':     company,
                'required':    required,
                'doc_number':  doc_number,
            })

        wb.close()
        systems = list(systems_seen.values())
        print(f"  Scope: {len(systems)} systems, {len(doc_types_ordered)} doc types, "
              f"{len(entries)} entries")
        return {'systems': systems, 'doc_types': doc_types_ordered, 'entries': entries}

    except Exception as e:
        print(f"  [warn] Scope parse error: {e}")
        return _mock_scope()


def _mock_scope():
    """Demo data covering a small chemical plant scope."""
    systems = [
        {'code':'CW-001', 'name':'Cooling Water',      'discipline':'Piping',         'company':'Piping AG',      'area':'Utility'},
        {'code':'RX-001', 'name':'Reactor Section',    'discipline':'Process Eng.',   'company':'Process GmbH',   'area':'Production'},
        {'code':'RX-002', 'name':'Reactor Effluent',   'discipline':'Process Eng.',   'company':'Process GmbH',   'area':'Production'},
        {'code':'HV-001', 'name':'HVAC Zone 1',        'discipline':'HVAC',           'company':'ClimaTech GmbH', 'area':'Building'},
        {'code':'EL-001', 'name':'MCC / Power Dist.',  'discipline':'Electrical',     'company':'Elektro AG',     'area':'Electrical Room'},
        {'code':'IN-001', 'name':'Instrumentation',    'discipline':'Instrumentation','company':'Instra GmbH',    'area':'Production'},
    ]
    doc_types = ['P&ID', 'Equipment List', 'Isometric', 'Loop Diagram', 'Layout Drawing', 'Datasheet']
    # Required matrix — not every doc type is required for every system
    required_map = {
        'CW-001': {'P&ID','Equipment List','Isometric','Layout Drawing'},
        'RX-001': {'P&ID','Equipment List','Isometric','Loop Diagram','Layout Drawing','Datasheet'},
        'RX-002': {'P&ID','Isometric','Loop Diagram','Datasheet'},
        'HV-001': {'Layout Drawing','Equipment List'},
        'EL-001': {'Loop Diagram','Layout Drawing','Datasheet'},
        'IN-001': {'Loop Diagram','Datasheet','Equipment List'},
    }
    entries = []
    for sys in systems:
        code = sys['code']
        req_set = required_map.get(code, set())
        for dt in doc_types:
            entries.append({
                'system_code': code,
                'doc_type':    dt,
                'discipline':  sys['discipline'],
                'company':     sys['company'],
                'required':    dt in req_set,
                'doc_number':  '',
            })
    return {'systems': systems, 'doc_types': doc_types, 'entries': entries}
