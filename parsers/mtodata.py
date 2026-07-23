"""MTO (Material Take-Off) data parser.

Output shape (matches existing JS exactly):
  {
    revisions: ['Rev A', 'Rev B', ...],   # ordered list of available revisions
    isometries: [{
      iso, tag, spool, dn, mat, pipeclass, discipline,
      revs: {'Rev A': {count, length_mm}, 'Rev B': {count, length_mm}}
    }],
    fittings: [{
      type, mat, dn,
      revs: {'Rev A': qty, 'Rev B': qty}
    }]
  }

config.yaml schema:
  mto:
    revisions:             # list of revision names in order
      - "Rev A"
      - "Rev B"
    isometries:
      path:  "./exports/mto/iso_register.xlsx"
      sheet: "Isometries"
      columns:
        iso_number:  "ISO Number"
        tag:         "Tag Number"
        spool:       "Spool Number"
        dn:          "DN"
        material:    "Material"
        pipeclass:   "Pipe Class"
        discipline:  "Discipline"
        # Revision columns: one per revision, key = revision name
        revisions:
          "Rev A":
            count:     "Rev A Count"
            length_mm: "Rev A Length (mm)"
          "Rev B":
            count:     "Rev B Count"
            length_mm: "Rev B Length (mm)"
    fittings:
      path:  "./exports/mto/fittings.xlsx"
      sheet: "Fittings"
      columns:
        type:      "Description"
        dn:        "DN"
        material:  "Material"
        revisions:
          "Rev A": "Qty Rev A"
          "Rev B": "Qty Rev B"
"""
from __future__ import annotations
from pathlib import Path


def parse_mto(cfg: dict) -> dict:
    mto_cfg = cfg.get('mto', {})
    if not mto_cfg:
        return _mock_mto()

    revisions  = mto_cfg.get('revisions', ['Rev A', 'Rev B'])
    isometries = _parse_isometries(mto_cfg.get('isometries', {}), revisions)
    fittings   = _parse_fittings(mto_cfg.get('fittings', {}), revisions)

    if not isometries and not fittings:
        return _mock_mto()

    print(f"  MTO: {len(isometries)} isometries, {len(fittings)} fittings, revisions: {revisions}")
    return {'revisions': revisions, 'isometries': isometries, 'fittings': fittings}


def _parse_isometries(iso_cfg: dict, revisions: list) -> list:
    if not iso_cfg or not iso_cfg.get('path'):
        return []
    path = Path(iso_cfg['path'])
    if not path.exists():
        print(f"  [warn] MTO isometries not found: {path}")
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb[iso_cfg['sheet']] if iso_cfg.get('sheet') in wb.sheetnames else wb.active
        cc = iso_cfg.get('columns', {})
        rev_cols = cc.get('revisions', {})  # {'Rev A': {'count': col, 'length_mm': col}}
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header: return []
        ci = {str(h).strip(): i for i, h in enumerate(header) if h}
        def gs(row, k):
            idx = ci.get(cc.get(k, k)); return str(row[idx]).strip() if idx is not None and row[idx] is not None else ''
        def gn(row, col_name, default=0):
            idx = ci.get(col_name); v = row[idx] if idx is not None and row[idx] is not None else default
            return float(v) if isinstance(v, (int, float)) else default
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            iso = gs(row, 'iso_number')
            if not iso: continue
            revs = {}
            for rev_name, rev_cc in rev_cols.items():
                revs[rev_name] = {
                    'count':     int(gn(row, rev_cc.get('count', ''), 0)),
                    'length_mm': int(gn(row, rev_cc.get('length_mm', ''), 0)),
                }
            rows.append({'iso': iso, 'tag': gs(row,'tag'), 'spool': gs(row,'spool'),
                         'dn': gs(row,'dn'), 'mat': gs(row,'material'),
                         'pipeclass': gs(row,'pipeclass'), 'discipline': gs(row,'discipline'),
                         'revs': revs})
        wb.close()
        return rows
    except Exception as e:
        print(f"  [warn] MTO isometries parse error: {e}")
        return []


def _parse_fittings(fit_cfg: dict, revisions: list) -> list:
    if not fit_cfg or not fit_cfg.get('path'):
        return []
    path = Path(fit_cfg['path'])
    if not path.exists():
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb[fit_cfg['sheet']] if fit_cfg.get('sheet') in wb.sheetnames else wb.active
        cc = fit_cfg.get('columns', {})
        rev_cols = cc.get('revisions', {})  # {'Rev A': col_name}
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header: return []
        ci = {str(h).strip(): i for i, h in enumerate(header) if h}
        def gs(row, k):
            idx = ci.get(cc.get(k, k)); return str(row[idx]).strip() if idx is not None and row[idx] is not None else ''
        def gn(row, col_name):
            idx = ci.get(col_name); v = row[idx] if idx is not None and row[idx] is not None else 0
            return float(v) if isinstance(v, (int, float)) else 0
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            ftype = gs(row, 'type')
            if not ftype: continue
            revs = {rev: int(gn(row, col)) for rev, col in rev_cols.items()}
            rows.append({'type': ftype, 'mat': gs(row,'material'), 'dn': gs(row,'dn'), 'revs': revs})
        wb.close()
        return rows
    except Exception as e:
        print(f"  [warn] MTO fittings parse error: {e}")
        return []


def _mock_mto():
    revisions = ['Rev A', 'Rev B', 'Rev C']
    isometries = [
        {'iso':'ISO-PP-001','tag':'1-H2SO4-100-A2B','spool':'SP-001','dn':'DN100','mat':'316L','pipeclass':'A2B','discipline':'Piping',
         'revs':{'Rev A':{'count':8,'length_mm':12400},'Rev B':{'count':9,'length_mm':13100},'Rev C':{'count':9,'length_mm':13100}}},
        {'iso':'ISO-PP-002','tag':'2-H2SO4-150-A2B','spool':'SP-002','dn':'DN150','mat':'316L','pipeclass':'A2B','discipline':'Piping',
         'revs':{'Rev A':{'count':6,'length_mm':8200},'Rev B':{'count':6,'length_mm':8200},'Rev C':{'count':7,'length_mm':9100}}},
        {'iso':'ISO-CW-001','tag':'3-H2O-200-C1','spool':'SP-003','dn':'DN200','mat':'CS','pipeclass':'C1','discipline':'Piping',
         'revs':{'Rev A':{'count':0,'length_mm':0},'Rev B':{'count':4,'length_mm':5600},'Rev C':{'count':4,'length_mm':5600}}},
        {'iso':'ISO-PP-003','tag':'5-TOL-80-B1','spool':'SP-004','dn':'DN80','mat':'CS','pipeclass':'B1','discipline':'Piping',
         'revs':{'Rev A':{'count':5,'length_mm':6800},'Rev B':{'count':5,'length_mm':6800},'Rev C':{'count':0,'length_mm':0}}},
        {'iso':'ISO-VT-001','tag':'7-N2-50-C1','spool':'SP-005','dn':'DN50','mat':'CS','pipeclass':'C1','discipline':'Piping',
         'revs':{'Rev A':{'count':3,'length_mm':3100},'Rev B':{'count':3,'length_mm':3100},'Rev C':{'count':3,'length_mm':3400}}},
    ]
    fittings = [
        {'type':'Elbow 90°','mat':'316L','dn':'DN100','revs':{'Rev A':12,'Rev B':14,'Rev C':14}},
        {'type':'Gate Valve','mat':'316L','dn':'DN80', 'revs':{'Rev A':6, 'Rev B':6, 'Rev C':5}},
        {'type':'Flange PN40','mat':'316L','dn':'DN150','revs':{'Rev A':24,'Rev B':28,'Rev C':28}},
        {'type':'Reducer','mat':'CS','dn':'DN200','revs':{'Rev A':0,'Rev B':4,'Rev C':4}},
    ]
    return {'revisions': revisions, 'isometries': isometries, 'fittings': fittings}
