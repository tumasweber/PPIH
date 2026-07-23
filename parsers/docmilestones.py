"""Document milestone / S-curve parser for the Document Management timeline chart.

config.yaml schema:
  doc_milestones:
    path:  "./exports/docregister/milestones.xlsx"
    sheet: "Milestones"
    # Each column after 'phase' is a dataset (discipline/source)
    columns:
      phase:  "Phase"        # x-axis labels
      # remaining columns become datasets automatically
"""
from __future__ import annotations
from pathlib import Path


PHASE_LABELS = ['Concept', 'Basic Design', 'Detail Design', 'IFC Issue', 'As-Built']
DATASET_COLORS = ['#73b5e2','#5ab87a','#e09a2a','#e05c5c','#9b59b6','#0081b1','#aab6be']


def parse_doc_milestones(cfg: dict) -> dict | None:
    ms_cfg = cfg.get('doc_milestones', {})
    if not ms_cfg or not ms_cfg.get('path'):
        return _mock_milestones()

    path = Path(ms_cfg['path'])
    if not path.exists():
        print(f"  [warn] Doc milestones not found: {path}")
        return _mock_milestones()

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb[ms_cfg['sheet']] if ms_cfg.get('sheet') in wb.sheetnames else wb.active
        cols_cfg = ms_cfg.get('columns', {})
        phase_col = cols_cfg.get('phase', 'Phase')

        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header: return _mock_milestones()
        headers = [str(h).strip() if h else '' for h in header]

        phase_idx = headers.index(phase_col) if phase_col in headers else 0
        data_cols = [(i, h) for i, h in enumerate(headers) if i != phase_idx and h]

        phases, datasets_raw = [], {name: [] for _, name in data_cols}
        for row in ws.iter_rows(min_row=2, values_only=True):
            phase = str(row[phase_idx] or '').strip()
            if not phase: continue
            phases.append(phase)
            for idx, name in data_cols:
                v = row[idx]
                datasets_raw[name].append(float(v) if v is not None else None)

        wb.close()
        datasets = [
            {'label': name, 'data': vals,
             'borderColor': DATASET_COLORS[i % len(DATASET_COLORS)],
             'backgroundColor': 'transparent', 'tension': 0.3,
             'spanGaps': True, 'pointRadius': 3}
            for i, (name, vals) in enumerate(datasets_raw.items())
        ]
        print(f"  Doc milestones: {len(phases)} phases, {len(datasets)} datasets")
        return {'labels': phases, 'datasets': datasets}
    except Exception as e:
        print(f"  [warn] Doc milestones parse error: {e}")
        return _mock_milestones()


def _mock_milestones():
    colors = DATASET_COLORS
    return {
        'labels': PHASE_LABELS,
        'datasets': [
            {'label':'SharePoint',   'data':[20,45,72,88,None],'borderColor':colors[0],'backgroundColor':'transparent','tension':0.3,'spanGaps':True,'pointRadius':3},
            {'label':'P&ID',         'data':[10,38,65,80,None],'borderColor':colors[1],'backgroundColor':'transparent','tension':0.3,'spanGaps':True,'pointRadius':3},
            {'label':'3D Model',     'data':[5, 25,58,75,None],'borderColor':colors[2],'backgroundColor':'transparent','tension':0.3,'spanGaps':True,'pointRadius':3},
        ]
    }
